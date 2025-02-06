from datetime import timedelta, timezone
from itertools import chain
import os
import re
import shutil
import subprocess
import time
from pathlib import Path
from openpyxl import load_workbook,Workbook
import pandas as pd
from pandas import DataFrame
from model import *
import random
from pymediainfo import MediaInfo
import pytz

def 取得所有檔案(path):
    """
    遍歷指定路徑下的所有檔案（包含子資料夾中的檔案），返回檔案清單。

    :param path: 路徑字串
    :return: 包含所有檔案完整路徑的清單
    """
    all_files = []
    for root, _, files in os.walk(path):  # `_` 忽略子資料夾名稱
        for file in files:
            all_files.append(os.path.join(root, file).replace("\\", "/"))
    return all_files

def 遍歷時間格式資料夾(base_path: str,year:bool = True) -> List[str]:
    """
    遍歷指定路徑下一階的資料夾，並篩選出符合民國年+月日格式 (如 1130101) 的資料夾，返回完整路徑列表。

    :param base_path: 要遍歷的基礎路徑
    :return: 符合條件的資料夾完整路徑列表
    """
    valid_folders = []
    
    # 民國年+月日的正則表達式
    if year:
        pattern = re.compile(r'^\d{3}(0[1-9]|1[0-2])(0[1-9]|[1-2]\d|3[0-1])$')
    else:
        pattern = re.compile(r'^(0[1-9]|1[0-2])(0[1-9]|[1-2]\d|3[0-1])$')
    
    try:
        # 遍歷 base_path 下的一階子項目
        for item in os.listdir(base_path):
            item_path = os.path.join(base_path, item)
            
            # 確保是資料夾且名稱符合格式
            if os.path.isdir(item_path) and pattern.match(item):
                valid_folders.append(item_path)
    except Exception as e:
        print(f"遍歷資料夾時出現錯誤: {str(e)}")
    
    return valid_folders

def 篩選包含民國年月日的路徑(path_list: List[str], 民國年月日: str) -> List[str]:
    """
    從路徑列表中篩選出包含指定民國年月日的路徑。

    :param path_list: 路徑的字符串列表
    :param 民國年月日: 民國年月日的字符串 (例如: 1120101)
    :return: 包含指定民國年月日的路徑列表
    """
    return [path for path in path_list if 民國年月日 in path]

def 查詢所有副檔名(files):
    # 取得副檔名並使用 set 去重複
    return list({os.path.splitext(file)[1] for file in files if '.' in file})

def 查詢檔案名稱包含關鍵字(路徑: str, 關鍵字: str) -> bool:
    # 儲存符合條件的檔案名稱
    符合條件檔案 = []
    
    # 確保路徑存在
    if not os.path.exists(路徑):
        print(F"查詢檔案名稱包含關鍵字 Error => 路徑:{路徑} 不存在")
        return False

    # 遍歷路徑下的所有檔案
    for 根目錄, _, 檔案名 in os.walk(路徑):
        for 檔案 in 檔案名:
            # 取得檔案名稱，不包含副檔名
            檔案名稱, _ = os.path.splitext(檔案)
            if 關鍵字 in 檔案名稱:
                # 如果檔案名稱包含關鍵字，加入結果列表
                符合條件檔案.append(os.path.join(根目錄, 檔案))
                print(F'{關鍵字}檔案已存在')
                return True
    return False

def 查詢檔案名稱包含關鍵字2(路徑: str, 關鍵字: str):
    
    # 確保路徑存在
    if not os.path.exists(路徑):
        print(F"查詢檔案名稱包含關鍵字 Error => 路徑:{路徑} 不存在")
        return None

    # 遍歷路徑下的所有檔案
    for 根目錄, _, 檔案名 in os.walk(路徑):
        for 檔案 in 檔案名:
            # 取得檔案名稱，不包含副檔名
            檔案名稱, _ = os.path.splitext(檔案)
            if 關鍵字 in 檔案名稱:
                查詢結果 = f'{根目錄}/{檔案}'
                print(F'{查詢結果}檔案已存在')
                return 查詢結果
    return None

def 查詢檔案名稱開頭指定字串(files, query:str):
    """
    從檔案列表中篩選出檔案名稱從第 0 個位置完全匹配查詢字串的檔案。
    :param files: 檔案完整路徑列表
    :param query: 查詢字串
    :return: 篩選後的檔案完整路徑列表
    """
    if query.isdigit():
        query_length = len(query)
        filtered_files = []
        for file in files:
            file_name = Path(file).name  # 取得檔案名稱（不含路徑）
            # 確保檔案名稱以查詢字串開頭
            if file_name[:query_length] == query:
                # 確認查詢字串後面不是數字
                if query_length == len(file_name) or not file_name[query_length].isalnum():
                    filtered_files.append(file)
        return filtered_files
    # 僅篩選檔案名稱以 query 完全匹配的檔案
    else:
        # 原版
        # filtered_files = [file for file in files if Path(file).name.lower()[:len(query)] == query.lower()]
        # 新版 
        # 修改此部分：忽略大小寫並僅匹配檔案名稱的第一部分
        filtered_files = []
        for file in files:
            file_name = Path(file).name.lower()
            # 檔案名按某個字符（如 "."）拆分，並檢查第一部分是否匹配
            if file_name.split('.')[0].startswith(query.lower()):
                filtered_files.append(file)
        return filtered_files
        
def 查詢檔案名稱指定字串(路徑列表:list, 檔名子字串:str)->list:
    """
    從路徑列表中查詢檔名符合指定子字串的檔案路徑。
    
    :param 路徑列表: list，包含檔案路徑的列表。
    :param 檔名子字串: str，需要匹配的檔名子字串。
    :return: list，符合條件的檔案路徑。
    """
    符合條件的檔案 = [
        路徑 for 路徑 in 路徑列表 if 檔名子字串 in Path(路徑).name
    ]
    return 符合條件的檔案    

def 提取檔名後段(file_path: str) -> str:
    """
    從檔案路徑提取檔名，去掉副檔名後，取得檔名中 '-' 後面的字串。

    :param file_path: 檔案路徑
    :return: 檔名中 '-' 後面的字串，如果沒有 '-' 則返回空字串
    """
    # 提取檔名（不含路徑）
    file_name_with_ext = os.path.basename(file_path)
    
    # 去掉副檔名
    file_name = os.path.splitext(file_name_with_ext)[0]
    
    # 找到 '-' 後面的部分
    if '-' in file_name:
        return file_name.split('-', 1)[1]
    else:
        return ""
    
def 提取檔名後段(file_path: str) -> str:
    """
    從檔案路徑提取檔名，去掉副檔名後，取得檔名中 '-' 後面的字串。

    :param file_path: 檔案路徑
    :return: 檔名中 '-' 後面的字串，如果沒有 '-' 則返回空字串
    """
    # 提取檔名（不含路徑）
    file_name_with_ext = os.path.basename(file_path)
    
    # 去掉副檔名
    file_name = os.path.splitext(file_name_with_ext)[0]
    
    # 找到 '-' 後面的部分
    if '-' in file_name:
        return file_name.split('-', 1)[1]
    else:
        return ""
    
def 提取檔名(file_path: str) -> str:
    """
    從檔案路徑提取檔名，去掉副檔名後，取得檔名中 '-' 後面的字串。

    :param file_path: 檔案路徑
    :return: 檔名中 '-' 後面的字串，如果沒有 '-' 則返回空字串
    """
    # 提取檔名（不含路徑）
    file_name_with_ext = os.path.basename(file_path)
    
    # 去掉副檔名
    return os.path.splitext(file_name_with_ext)[0]
    

def 提取路徑時間並轉換(file_path: str,西元: bool = True) -> datetime:
    """
    從路徑中提取時間部分（民國年月日格式），並轉換為 datetime 對象。

    :param file_path: 檔案的完整路徑
    :return: 對應的 datetime 對象，如果路徑中不包含有效的時間則拋出 ValueError
    """
    import re

    if 西元:
        # 匹配西元年月日格式，例如 20240103
        match = re.search(r'(\d{4})(\d{2})(\d{2})', file_path)
        if not match:
            raise ValueError("路徑中沒有找到有效的西元年月日格式")
        # 提取西元年、月、日
        year, month, day = map(int, match.groups())
    else:
        # 匹配民國年月日格式，例如 1130103
        match = re.search(r'(\d{3})(\d{2})(\d{2})', file_path)
        if not match:
            raise ValueError("路徑中沒有找到有效的民國年月日格式")
        # 提取民國年、月、日
        roc_year, month, day = map(int, match.groups())
        # 將民國年轉為西元年
        year = roc_year + 1911

    # 建立 datetime 對象
    return datetime(year, month, day)

def 提取路徑中的月日(path: str) -> str:
    """
    從路徑中提取符合 MMDD 格式的部分，並確保只提取資料夾名稱中的內容，且完全匹配。

    :param path: 路徑字串
    :return: 符合 MMDD 格式的資料夾名稱列表
    """
    # 定義 MMDD 的正則表達式
    pattern = r'^(0[1-9]|1[0-2])(0[1-9]|[1-2]\d|3[0-1])$'
    
    # 分割路徑為資料夾名稱
    path_parts = path.split(os.altsep)
    
    # 遍歷路徑中的每一部分，找到完全符合的 MMDD 格式
    mmdd_list = [part for part in path_parts if re.match(pattern, part)]
    
    return mmdd_list[0]

def 過濾全字串關鍵字(input_list:list, keywords:list):
    """
    過濾掉包含任意關鍵字的元素，返回剩餘的列表。

    :param input_list: List[str] - 要篩選的字串列表
    :param keywords: List[str] - 關鍵字列表
    :return: List[str] - 過濾後的字串列表
    """
    return [item for item in input_list if not any(keyword in item for keyword in keywords)]

def 查詢全路徑關鍵字(input_list:list, keywords:list):
    """
    過濾掉包含任意關鍵字的元素，返回剩餘的列表。

    :param input_list: List[str] - 要篩選的字串列表
    :param keywords: List[str] - 關鍵字列表
    :return: List[str] - 過濾後的字串列表
    """
    return [item for item in input_list if not any(keyword in item for keyword in keywords)]

def 複製檔案(source_path, destination_path):
    """
    複製檔案到指定的目標路徑。
    :param source_path: 原始檔案路徑
    :param destination_path: 複製到的目標路徑
    """
    try:
        # 確保目標資料夾存在
        os.makedirs(os.path.dirname(destination_path), exist_ok=True)
        shutil.copy2(source_path, destination_path)
        print(f"檔案已成功複製到: {destination_path}")
    except Exception as e:
        print(f"複製檔案失敗: {e}")

def 修改影片日期(file_path, new_date):
    """
    修改影片的最後修改日期和訪問日期為指定時間戳。
    :param file_path: 影片檔案路徑
    :param new_date: 新的日期 (格式: "YYYY-MM-DD HH:MM:SS")
    """
    try:
        # 將日期轉為時間戳
        timestamp = time.mktime(time.strptime(new_date, "%Y-%m-%d %H:%M:%S"))
        # 修改檔案的訪問和修改日期
        os.utime(file_path, (timestamp, timestamp))
        print(f"{file_path} 的日期已更新為 {new_date}")
    except Exception as e:
        print(f"修改檔案日期失敗: {e}")

def 讀取excel檔案(file_path):
    """
    读取 .xlsx 文件并返回其内容
    :param file_path: 文件路径
    :return: 字典，包含工作表名称和对应内容
    """
    try:
        # 加载工作簿
        workbook:Workbook = load_workbook(file_path)
        
        # 存储结果的字典
        data = {}
        
        # 遍历每个工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            
            # 遍历每一行
            for row in sheet.iter_rows(values_only=True):
                sheet_data.append(row)
            
            # 将工作表数据存储到字典
            data[sheet_name] = sheet_data
        
        return data
    except Exception as e:
        print(f"读取文件时出错：{e}")
        return None
    
def 讀取並清洗數據(file_path)->DataFrame:
    """
    读取 Excel 文件并清洗数据，返回清洗后的 DataFrame。
    """
    # 读取 Excel 文件，跳过最上面两行
    df:DataFrame = pd.read_excel(file_path, skiprows=2)

    # 删除全空行
    df = df.dropna(how='all')

    # 确保巡查日期是 datetime 格式
    df['巡查日期'] = pd.to_datetime(df['巡查日期'], format="%m/%d/%Y", errors='coerce')
    df['靠離時間'] = pd.to_datetime(df['靠離時間'], errors='coerce')

    return df

def 讀取csv並清洗數據(file_path)->DataFrame:
    """
    读取 Excel 文件并清洗数据，返回清洗后的 DataFrame。
    """
    # 读取 Excel 文件，跳过最上面两行
    df:DataFrame = pd.read_csv(file_path, skiprows=2)

    # 删除全空行
    df = df.dropna(how='all')

    # 确保巡查日期是 datetime 格式
    df['巡查日期'] = pd.to_datetime(df['巡查日期'], format="%Y/%m/%d", errors='coerce')
    df['靠離時間'] = pd.to_datetime(df['靠離時間'], errors='coerce')

    return df

def 查詢日期(dataframe:DataFrame, month:int)-> DataFrame:
    """
    根据指定的月份筛选巡查数据。
    :param dataframe: 清洗后的 DataFrame
    :param month: 要筛选的月份（整数，1-12）
    :return: 筛选后的 DataFrame
    """
    return dataframe[dataframe['巡查日期'].dt.month == month]

def 查詢影片建立日期(file_path):
    try:
        # 獲取檔案建立時間（UNIX 時間戳）
        creation_time = os.path.getctime(file_path)
        # 格式化為可讀的日期時間
        creation_date = datetime.fromtimestamp(creation_time)
        return creation_date
    except Exception as e:
        print(f"Error: {e}")
        return None

def 查詢影片媒體建立時間_mediainfo(video_path):
    try:
        media_info: MediaInfo = MediaInfo.parse(video_path)
        for track in media_info.tracks:
            if track.track_type == "General":
                creation_time = track.tagged_date
                if creation_time:
                    # 保持原來時區，不強制轉換
                    creation_time = creation_time.replace(" UTC", "")
                    媒體建立日期 = datetime.fromisoformat(creation_time)
                    print(媒體建立日期.strftime("%Y-%m-%d %H:%M:%S"))
                    return 媒體建立日期
        # 如果没有找到媒体创建时间
        return datetime(1900, 1, 1, 0, 0)
    except Exception as e:
        # 发生错误时返回默认时间
        print(f"查詢影片媒體建立時間时发生错误: {e}")
        return datetime(1900, 1, 1, 0, 0)

def 修改特定日期(file_path, new_date:datetime):
    try:
        # 自动检测系统本地时区
        local_timezone = datetime.now(timezone.utc).astimezone().tzinfo
        new_date = new_date.replace(tzinfo=local_timezone)
        # 將日期轉換為 UNIX 時間戳
        timestamp = new_date.timestamp()
        # 獲取目前訪問時間（保留原訪問時間）
        current_access_time = os.stat(file_path).st_atime
        # 使用 os.utime 更新檔案時間
        os.utime(file_path, (current_access_time, timestamp))
        print(f"已更新修改日期為: {new_date}")
    except Exception as e:
        print(f"Error: {e}")

def 轉換為民國年月日(dt: datetime) -> str:
    """
    將 datetime 轉換為民國年月日格式的字符串 (例如: 1120101)。

    :param dt: 要轉換的 datetime 對象
    :return: 民國年月日格式的字符串
    """
    民國年 = dt.year - 1911
    return f"{民國年:03d}{dt.month:02d}{dt.day:02d}"

def 轉換為月日(dt: datetime) -> str:
    return f"{dt.month:02d}{dt.day:02d}"

def 複製並重新命名檔案(source_file, destination_file):
    try:
        # 確保目標資料夾存在
        destination_folder = os.path.dirname(destination_file)
        os.makedirs(destination_folder, exist_ok=True)
        
        # 複製並重新命名檔案
        shutil.copy(source_file, destination_file)
        print(f"檔案已成功複製並重新命名為: {destination_file}")
    except Exception as e:
        print(f"複製並重新命名檔案時發生錯誤: {e}")

def 截斷號碼頭(碼頭列表):
    """
    從碼頭列表中移除「號碼頭」字串。
    """
    return [碼頭.replace('號碼頭', '') for 碼頭 in 碼頭列表]

def 隨機減少分鐘(輸入時間:datetime, 隨機分鐘數:int)->datetime:
    """
    隨機減少指定的分鐘數範圍內的時間，並返回新的 datetime 時間。

    :param original_datetime: datetime，原始時間
    :param max_minutes: int，最大減少的分鐘數
    :return: datetime，新的時間
    """
    if 隨機分鐘數 < 0:
        raise ValueError("max_minutes 必須是非負整數")

    # 隨機生成 0 到 max_minutes 的減少分鐘數
    random_minutes = random.randint(20, 隨機分鐘數)

    # 減少時間
    new_datetime = 輸入時間 - timedelta(minutes=random_minutes)
    return new_datetime

def 判斷白天或晚上(輸入時間:datetime, 白天起:datetime, 白天迄:datetime)->str:
    """
    判斷輸入的時間是否在白天區間內，返回 '白天' 或 '晚上'。

    :param input_datetime: datetime，輸入的時間
    :param daytime_start: datetime，白天起始時間
    :param daytime_end: datetime，白天結束時間
    :return: str，'白天' 或 '晚上'
    """
    # 判斷是否在白天區間內
    if 白天起.time() <= 輸入時間.time() <= 白天迄.time():
        return '白天'
    else:
        return '晚上'

def 判斷白天或晚上2(時間: datetime) -> str:
    if 6 <= 時間.hour < 18: # 早上6.~晚上6.為白天
        return "白天"
    else:
        return "晚上"

def 隨機選擇符合時間的影片(files: list,天空:str):
    # 記錄最多嘗試次數
    最大嘗試次數 = 10
    嘗試次數 = 0
    try:
        while 嘗試次數 < 最大嘗試次數:
            # 隨機選擇影片
            隨機影片 = random.choice(files)
            媒體建立日期 = 查詢影片媒體建立時間_mediainfo(隨機影片)
            媒體天空= 判斷白天或晚上2(媒體建立日期)
            
            if 媒體建立日期 != datetime(1900, 1, 1, 0, 0):  # 確保媒體建立時間有效
                # 判斷該影片的建立時間是白天還是晚上
                if 天空 == 媒體天空:
                    # print(f"影片 {隨機影片} 的建立時間為白天 ({媒體建立日期.strftime('%Y-%m-%d %H:%M:%S')})")
                    return 隨機影片
            嘗試次數 += 1

        # 如果10次都沒有符合條件的影片，遍歷所有影片
        for 影片 in files:
            媒體建立日期 = 查詢影片媒體建立時間_mediainfo(影片)
            媒體天空 = 判斷白天或晚上2(媒體建立日期)
            if 媒體建立日期 != datetime(1900, 1, 1, 0, 0):  # 確保媒體建立時間有效
                if 天空 == 媒體天空:
                    # print(f"影片 {影片} 的建立時間為白天 ({媒體建立日期.strftime('%Y-%m-%d %H:%M:%S')})")
                    return 影片

        # 如果完全沒有符合的情況，隨機返回一部影片
        隨機影片 = random.choice(files)
        print(f"沒有符合條件的影片，隨機返回影片 {隨機影片}")
        return 隨機影片
    except Exception as e:
        print(f'隨機選擇符合時間的影片 error => {e.args[0]}')
        return None

def 創建檔案(檔案路徑: str):
    # 確保路徑存在，若不存在則創建
    目錄路徑 = os.path.dirname(檔案路徑)
    if not os.path.exists(目錄路徑):
        os.makedirs(目錄路徑)
    
    # 使用 'w' 模式開啟檔案，若檔案不存在會自動創建
    try:
        with open(檔案路徑, 'w') as file:
            # 這裡可以寫入初始化內容，若不需要可以省略
            file.write("")  # 如果要創建空檔案，這行可以保留
        print(f"檔案已創建: {檔案路徑}")
    except Exception as e:
        print(f"創建檔案時發生錯誤: {e}")

def 修改影片元數據建立日期(file_path: str, 輸入時間: datetime):
    """
    修改影片檔案的 CreateDate 屬性，並直接覆蓋原始檔案。

    :param file_path: 影片檔案的路徑
    :param new_create_date: 要設置的新的創建日期，datetime 格式
    :return: 修改結果
    """
    try:
        # 將 datetime 物件轉換為 'YYYY:MM:DD HH:MM:SS' 格式的字符串
        local_tz = pytz.timezone("Asia/Taipei")  # 設定時區為台北時間，根據需求修改
        輸入時間_時區 = local_tz.localize(輸入時間,is_dst=True)  # 將 datetime 設為台北時區時間
        輸入時間_時區_str = 輸入時間_時區.strftime('%Y:%m:%d %H:%M:%S')
        輸入時間_utc_str = 輸入時間_時區.astimezone(pytz.utc).strftime('%Y:%m:%d %H:%M:%S')
        
        # 構造 exiftool 命令，添加 -overwrite_original 以覆蓋原始檔案
        # command = ['exiftool', f'-CreateDate={new_create_date_str}', '-overwrite_original', file_path]
        command = ['C:\\exiftool\\exiftool-13.11_64\\exiftool.exe', f'-CreateDate={輸入時間_utc_str}', '-overwrite_original', file_path]
        
        # 執行命令
        result = subprocess.run(command, capture_output=True, text=True, shell=True)
        
        # 檢查是否成功
        if result.returncode == 0:
            print(f"成功 修改影片元數據建立日期 => {file_path} 的 媒體建立日期 為 {輸入時間_時區_str}")
        else:
            print(f"失敗 修改影片元數據建立日期 => {result.stderr}")
    except Exception as e:
        print(f"異常 修改影片元數據建立日期 => {str(e)}")

def main():
    #修改檔案媒體日期
    目標路徑 = 'D:/sister/薩迦曼丹.MOV'

    一到十月檔案資訊 = 讀取csv並清洗數據('./113年1-10月靠離岸巡查2.csv')
    舊船隻資訊列表 = [轉換為船隻資訊_新格式(row) for index, row in 一到十月檔案資訊.iterrows()]
    所有船隻資訊 = 舊船隻資訊列表
    船名 = 提取檔名(目標路徑)
    # 檔案月日 = 提取路徑中的月日(目標路徑)
    # 檔案時間 = 提取路徑時間並轉換(目標路徑)
    檔案時間 = datetime(year=2024,month=6,day=2)
    船隻資訊 = [船隻 for 船隻 in 所有船隻資訊 if 轉換為月日(船隻.巡查日期) == 轉換為月日(檔案時間) and 船名 in 船隻.船名]
    for 船隻 in 船隻資訊:
        print(f'路徑檔案:{目標路徑}')
        print(f'正在處理 {船隻.靠離時間} {船隻.設施名稱} {船隻.船名}')
        靠離時間 = 船隻.靠離時間
        假設拍攝時間 = 隨機減少分鐘(靠離時間,110)
        修改影片元數據建立日期(目標路徑,假設拍攝時間)
    print('修改完成')     

if __name__ == "__main__":
    main()